#!/usr/bin/env python3
"""
qa.py — Data Quality tool for ITS Infrastructure Health Monitor.

Reads from the local SQLite DB and optionally one or more reference CSV/Excel
files, then writes a self-contained HTML report with a Leaflet map.

Usage:
    python runner/qa.py --group VMS
    python runner/qa.py --group VMS --ref large.csv b_type.xlsx
    python runner/qa.py --group VMS --ref large.csv --out reports/qa_vms.html

Reference files must have at minimum a 'name' column and a 'location' column.
The location column must contain DMS coordinates, e.g.:
    34° 42' 53.4488" N 32° 32' 39.458" E
    34°41'32.1"N 32°58'31.3"E
Any extra columns (district, description, type, notes) are carried through as
metadata and shown in the report.
"""

import argparse
import html
import json
import math
import os
import re
import sys
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
os.environ.setdefault("DB_PATH", str(Path(__file__).parent.parent / "results" / "history.db"))

# Load .env from repo root if present
_env_path = Path(__file__).parent.parent / ".env"
if _env_path.exists():
    for _line in _env_path.read_text(encoding='utf-8').splitlines():
        _line = _line.strip()
        if _line and not _line.startswith('#') and '=' in _line:
            _k, _, _v = _line.partition('=')
            os.environ.setdefault(_k.strip(), _v.strip())
from db import get_connection, upsert_sensor_projects
from stability import HEALTH_GOOD_PCT, HEALTH_WARNING_PCT

REPORT_DIR      = Path(__file__).parent.parent / "reports"
MATCH_THRESHOLD = 0.60   # minimum similarity to attempt a match
COORD_WARN_M    = 100    # metres — flag yellow above this
COORD_ALERT_M   = 500    # metres — flag red above this

PROJECTS_CSV = Path(__file__).parent.parent / "config" / "projects.csv"
COLOCATION_M = 15   # API detectors this close are treated as the same installation
BONUS_HEALTH_PCT = 50   # out-of-support sensor counts as genuine "bonus" only at/above this uptime

# Accountability — who is on the hook when a sensor fails.
#   supported        → in-contract; a failure is actionable (normal red/amber/green)
#   out_of_support   → contract ended; a failure is expected (non-actionable), a
#                      working sensor is bonus value. Excluded from health stats.
#   not_commissioned → installed but not yet live; not expected to work yet.
#
# A sensor's project comes from the REFERENCE spreadsheet (the equipment list we
# were handed), matched to the API sensor by location. projects.csv then maps that
# project to its accountability. We deliberately do NOT derive project from any
# API-sourced data — that would be validating the API against itself.
ACCT_SUPPORTED        = "supported"
ACCT_OUT_OF_SUPPORT   = "out_of_support"
ACCT_NOT_COMMISSIONED = "not_commissioned"


def load_project_accountability():
    """Return {project_name: accountability} from config/projects.csv."""
    import csv as _csv
    status = {}
    if PROJECTS_CSV.exists():
        with open(PROJECTS_CSV, newline='', encoding='utf-8-sig') as f:
            for r in _csv.DictReader(f):
                proj = (r.get('project') or '').strip()
                acct = (r.get('accountability') or '').strip().lower()
                if proj:
                    status[proj] = acct or ACCT_SUPPORTED
    return status


def annotate_accountability(api_sensors, matches, project_acct, max_dist=None):
    """Attach 'project', 'accountability', and 'project_source' to each API sensor.

    Project is taken from the matched reference sensor's 'project' column;
    accountability is looked up from projects.csv. Then, because the API often
    reports two co-located detectors (e.g. TCC + ACC) at one physical site while
    the reference lists it once, the project is propagated to any *unmatched* API
    sensor within COLOCATION_M of a matched one — two detectors at the same point
    are the same installation, hence the same contract.

    project_source records how the project was determined:
      'matched'   — directly matched to a reference row
      'colocated' — inherited from a co-located matched twin

    When no project is found, project_source records *why*, so the dashboard can
    tell a missing spreadsheet row apart from a row that a nearer sensor claimed:
      'unmatched_no_ref:<m>'     — nearest reference row is <m> metres away, beyond
                                   the match threshold; the row does not exist
      'unmatched_ref_taken:<m>'  — a reference row sits <m> metres away, inside the
                                   threshold, but a closer sensor already claimed it
      'unmatched_no_coords'      — the sensor has no coordinates to match on
    """
    threshold = max_dist if max_dist is not None else COORD_MATCH_MAX_M
    for s in api_sensors:
        s.setdefault('project', None)
        s.setdefault('accountability', ACCT_SUPPORTED)
        s.setdefault('project_source', None)
        s.setdefault('commissioning', 'active')

    matched = []
    for m in matches:
        if m['type'] != 'match':
            continue
        proj = (m['ref'].get('extra', {}).get('project') or '').strip() or None
        api = m['api']
        api['project']        = proj
        api['accountability']  = project_acct.get(proj, ACCT_SUPPORTED) if proj else ACCT_SUPPORTED
        api['project_source']  = 'matched'
        api['commissioning']   = m['ref'].get('commissioning', 'active')
        matched.append(api)

    # Co-location propagation for unmatched sensors
    matched_ids = {s['id'] for s in matched}
    donors = [m for m in matched if m['project'] and m['lat'] is not None]
    for s in api_sensors:
        if s['id'] in matched_ids or s['project_source'] is not None or s['lat'] is None:
            continue
        twin = min(
            (m for m in donors if _haversine_m(s['lat'], s['lon'], m['lat'], m['lon']) <= COLOCATION_M),
            key=lambda m: _haversine_m(s['lat'], s['lon'], m['lat'], m['lon']),
            default=None,
        )
        if twin:
            s['project']        = twin['project']
            s['accountability'] = twin['accountability']
            s['project_source'] = 'colocated'
            s['commissioning']  = twin.get('commissioning', 'active')

    # Record why the stragglers went unassigned. Every reference row appears in
    # matches as either a 'match' or a 'ref_only', so this is the full sheet.
    all_refs = [m['ref'] for m in matches
                if m.get('ref') and m['ref'].get('lat') is not None]
    for s in api_sensors:
        if s['project_source'] is not None:
            continue
        if s['lat'] is None:
            s['project_source'] = 'unmatched_no_coords'
            continue
        if not all_refs:
            s['project_source'] = 'unmatched_no_ref:'
            continue
        dist = min(_haversine_m(s['lat'], s['lon'], r['lat'], r['lon']) for r in all_refs)
        kind = 'unmatched_ref_taken' if dist <= threshold else 'unmatched_no_ref'
        s['project_source'] = f'{kind}:{round(dist)}'
    return api_sensors


def annotate_ref_accountability(ref_sensors, project_acct):
    """Attach 'project' + 'accountability' to each reference sensor from its own
    Project column. Lets out-of-support reference locations (e.g. a Diavlos gap
    the API no longer reports) be flagged as expected rather than a concern."""
    for r in ref_sensors:
        proj = (r.get('extra', {}).get('project') or '').strip() or None
        r['project']        = proj
        r['accountability'] = project_acct.get(proj, ACCT_SUPPORTED) if proj else ACCT_SUPPORTED
    return ref_sensors


# ── Coordinate helpers ────────────────────────────────────────────────────────

def _dms_to_dd(text):
    """Parse a DMS string to (lat, lon) decimal degrees. Returns (None, None) on failure."""
    if not text or not isinstance(text, str):
        return None, None
    # [°°�] covers proper °, and common corruption variants
    pat = (r'(\d+)\s*[°°�]\s*(\d+)\s*[\'′’]\s*([\d.]+)\s*["""”]\s*([NSns])'
           r'\s+'
           r'(\d+)\s*[°°�]\s*(\d+)\s*[\'′’]\s*([\d.]+)\s*["""”]\s*([EWew])')
    m = re.search(pat, text.strip())
    if not m:
        return None, None
    d1, m1, s1, h1, d2, m2, s2, h2 = m.groups()
    lat = int(d1) + int(m1) / 60 + float(s1) / 3600
    if h1.upper() == 'S':
        lat = -lat
    lon = int(d2) + int(m2) / 60 + float(s2) / 3600
    if h2.upper() == 'W':
        lon = -lon
    return round(lat, 7), round(lon, 7)


def _kml_to_dd(text):
    """Parse a KML 'lon,lat' string to (lat, lon). Returns (None, None) on failure."""
    if not text or not isinstance(text, str):
        return None, None
    parts = text.strip().split(',')
    if len(parts) >= 2:
        try:
            lon, lat = float(parts[0].strip()), float(parts[1].strip())
            if -90 <= lat <= 90 and -180 <= lon <= 180:
                return round(lat, 7), round(lon, 7)
        except (ValueError, TypeError):
            pass
    return None, None


def _haversine_m(lat1, lon1, lat2, lon2):
    """Distance in metres between two lat/lon points."""
    R = 6_371_000
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlam / 2) ** 2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


# ── Reference file loader ─────────────────────────────────────────────────────

# Ordered by priority — higher-specificity names first so 'location gps' wins over 'location' or 'column1'
_LOCATION_COLS = ['location gps', 'gps location', 'gps', 'kml coordinates', 'kml', 'coordinates', 'coords', 'location', 'column1']
_LAT_COLS      = ['latitude', 'lat']
_LON_COLS      = ['longitude', 'lon', 'long']
# Prefer descriptive names ('name'/'vms X') over an opaque 'key' code; 'key' stays
# as a per-row fallback (see _parse_rows) so a blank name still yields something.
_NAME_COLS     = ['vms b', 'vms a', 'vms c', 'vms', 'sensor name', 'name', 'key', 'sensor', 'id']


def _detect_col(headers, candidates):
    """Return the header that matches the highest-priority candidate (case-insensitive)."""
    header_lower = {h.strip().lower(): h for h in headers}
    for candidate in candidates:
        if candidate in header_lower:
            return header_lower[candidate]
    return None


_INACTIVE_VALUES = {'inactive', 'removed', 'decommissioned', 'retired', 'no', 'false', '0', 'off'}

def _detect_cols(headers, candidates):
    """Return every header matching a candidate, in candidate-priority order."""
    header_lower = {h.strip().lower(): h for h in headers}
    return [header_lower[c] for c in candidates if c in header_lower]


_NOT_ELECTRIFIED = {'not electrified', 'pending power'}


def _parse_rows(raw_rows):
    """Given list of dicts with lowercased keys, extract name/location/extras.

    Returns (sensors, not_electrified_count) — the count is rows skipped
    specifically because their status was 'not electrified' (installed but
    not yet commissioned), tracked separately from other inactive statuses
    so the report can call it out explicitly.
    """
    if not raw_rows:
        return [], 0
    headers = list(raw_rows[0].keys())
    # A file may have several name-like columns (e.g. 'key' AND 'name'); pick the
    # first non-empty one per row so a blank cell in the top-priority column
    # doesn't drop the whole row (e.g. BT rows with a blank 'key' but a real 'name').
    name_cols  = _detect_cols(headers, _NAME_COLS)
    loc_col    = _detect_col(headers, _LOCATION_COLS)
    lat_col    = _detect_col(headers, _LAT_COLS)
    lon_col    = _detect_col(headers, _LON_COLS)
    status_col = _detect_col(headers, ['status', 'active', 'state'])
    used_cols  = set(name_cols) | {c for c in (loc_col, lat_col, lon_col, status_col) if c}
    sensors = []
    not_electrified = 0
    for row in raw_rows:
        name = ''
        for nc in name_cols:
            val = str(row.get(nc, '') or '').strip()
            if val and val.lower() not in ('name', 'vms', 'key', 'nan', 'none'):
                name = val
                break
        if not name:
            continue
        commissioning = 'active'
        if status_col:
            status_val = row.get(status_col, '').strip().lower()
            # Not-electrified and inactive rows are both KEPT (so they still match
            # their API sensor and carry the state through), just flagged — the
            # dashboard excludes both from health stats. Substring match for
            # not-electrified so "pending power" catches "pending power connection";
            # _INACTIVE_VALUES stays exact (its values are short).
            if any(k in status_val for k in _NOT_ELECTRIFIED):
                not_electrified += 1
                commissioning = 'not_electrified'
            elif status_val in _INACTIVE_VALUES:
                commissioning = 'decommissioned'
        # Prefer separate lat/lon columns over combined DMS column
        if lat_col and lon_col:
            try:
                lat = float(row.get(lat_col, '') or '')
                lon = float(row.get(lon_col, '') or '')
            except (ValueError, TypeError):
                lat, lon = None, None
            loc_raw = f'{lat}, {lon}' if lat is not None else ''
        else:
            loc_raw = row.get(loc_col, '').strip() if loc_col else ''
            lat, lon = _dms_to_dd(loc_raw)
            if lat is None:
                lat, lon = _kml_to_dd(loc_raw)
        extra = {k: v for k, v in row.items()
                 if k not in used_cols and v and str(v).lower() not in ('nan', 'none', '')}
        sensors.append({
            'name': name, 'lat': lat, 'lon': lon,
            'loc_raw': loc_raw, 'extra': extra, 'commissioning': commissioning,
        })
    return sensors, not_electrified


def _load_csv(path):
    import csv
    with open(path, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        return [{k.strip().lower(): (v or '').strip() for k, v in row.items()}
                for row in reader]


def _load_excel(path, sheet_name=None):
    try:
        import openpyxl
    except ImportError:
        raise SystemExit(
            f"openpyxl is required to read Excel files.\n"
            f"Install with:  pip install openpyxl\n"
            f"Or convert {Path(path).name} to CSV first."
        )
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise SystemExit(f"ERROR: sheet '{sheet_name}' not found in {Path(path).name}. "
                              f"Available sheets: {', '.join(wb.sheetnames)}")
        ws = wb[sheet_name]
    else:
        ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower() if h is not None else '' for h in rows[0]]
    return [
        {headers[i]: str(v).strip() if v is not None else '' for i, v in enumerate(row)}
        for row in rows[1:]
    ]


def load_reference(paths):
    """Load CSV/Excel reference files. Returns list of sensor dicts with source filename.

    A path may include '::SheetName' to select a specific sheet from a
    multi-sheet workbook, e.g. 'QA Locations.xlsx::Bluetooth'.
    """
    sensors = []
    not_electrified = 0
    for p in paths:
        p = str(p)
        sheet_name = None
        if '::' in p:
            p, sheet_name = p.split('::', 1)
        p = Path(p)
        raw = _load_excel(p, sheet_name) if p.suffix.lower() in ('.xlsx', '.xls') else _load_csv(p)
        parsed, ne = _parse_rows(raw)
        not_electrified += ne
        source_label = f"{p.name} [{sheet_name}]" if sheet_name else p.name
        for s in parsed:
            s['source'] = source_label
        sensors.extend(parsed)
        msg = f"  {source_label}: {len(parsed)} sensors loaded"
        if ne:
            msg += f" ({ne} awaiting power — flagged, kept for matching)"
        print(msg)
    return sensors, not_electrified


# ── DB reader ─────────────────────────────────────────────────────────────────

def load_api_sensors(group_name):
    """Load sensors from DB for the given group. Returns list of dicts."""
    conn = get_connection()
    coords = conn.execute(
        "SELECT sensor_id, name, lat, lon, active, last_seen "
        "FROM sensor_coords WHERE group_name=?",
        (group_name,)
    ).fetchall()

    stability = conn.execute("""
        SELECT sensor_id,
               COUNT(*)  AS total_runs,
               SUM(CASE WHEN status IN ('working', 'ok') THEN 1 ELSE 0 END) AS good_runs,
               MIN(run_at) AS first_seen,
               MAX(run_at) AS last_run
        FROM sensor_results
        WHERE group_name=?
        GROUP BY sensor_id
    """, (group_name,)).fetchall()
    conn.close()

    health = {r['sensor_id']: dict(r) for r in stability}
    sensors = []
    for r in coords:
        sid = r['sensor_id']
        h   = health.get(sid, {})
        total = h.get('total_runs', 0)
        good  = h.get('good_runs', 0)
        pct   = round(good / total * 100, 1) if total > 0 else None
        sensors.append({
            'id':         sid,
            'name':       r['name'] or sid,
            'lat':        r['lat'],
            'lon':        r['lon'],
            'active':     bool(r['active']),
            'last_seen':  r['last_seen'],
            'total_runs': total,
            'health_pct': pct,
            'first_seen': h.get('first_seen'),
            'last_run':   h.get('last_run'),
        })
    return sensors


# ── Live API fetch ────────────────────────────────────────────────────────────

def load_api_sensors_live(group_name):
    """Fetch VmsTablePublication live and return sensors in the same format as load_api_sensors."""
    import httpx
    from geo import extract_vms_coords, extract_measurement_site_coords

    base_url = os.environ.get('BASE_URL', '').strip().rstrip('/')
    swarco   = os.environ.get('SWARCO', '').strip().strip('/')
    if not base_url or not swarco:
        raise SystemExit(
            "ERROR: --live requires BASE_URL and SWARCO environment variables to be set.\n"
            "       Set them in your .env or environment, then re-run."
        )

    # Pick the right endpoint and parser based on group name
    group_lower = group_name.lower()
    if 'vms' in group_lower:
        path, parser = 'VmsTablePublication', extract_vms_coords
    elif 'bluetooth' in group_lower or 'bt' in group_lower:
        path, parser = 'BTMeasurementSiteTablePublication', extract_measurement_site_coords
    else:
        path, parser = 'TrafficMeasurementSiteTablePublication', extract_measurement_site_coords

    url = f"{base_url}/{swarco}/{path}"
    print(f"  Fetching {url} ...")
    try:
        import time
        t0 = time.perf_counter()
        resp = httpx.get(url, timeout=30, follow_redirects=True)
        elapsed = (time.perf_counter() - t0) * 1000
        resp.raise_for_status()
        print(f"  HTTP {resp.status_code}  {elapsed:.0f} ms")
    except Exception as e:
        raise SystemExit(f"ERROR: Live fetch failed: {e}")

    coords = parser(resp.text)
    if not coords:
        raise SystemExit("ERROR: Live fetch returned no sensors — check the API response.")
    print(f"  Live data fetched OK — {len(coords)} sensors from API")

    # Load health from DB (coords are live, history still comes from DB)
    conn = get_connection()
    stability = conn.execute("""
        SELECT sensor_id,
               COUNT(*)  AS total_runs,
               SUM(CASE WHEN status IN ('working', 'ok') THEN 1 ELSE 0 END) AS good_runs
        FROM sensor_results WHERE group_name=? GROUP BY sensor_id
    """, (group_name,)).fetchall()
    conn.close()
    health = {r['sensor_id']: dict(r) for r in stability}

    sensors = []
    for sid, c in coords.items():
        h     = health.get(sid, {})
        total = h.get('total_runs', 0)
        good  = h.get('good_runs', 0)
        pct   = round(good / total * 100, 1) if total > 0 else None
        sensors.append({
            'id':         sid,
            'name':       c.get('name') or sid,
            'lat':        c.get('lat'),
            'lon':        c.get('lon'),
            'active':     True,
            'last_seen':  None,
            'total_runs': total,
            'health_pct': pct,
            'first_seen': None,
            'last_run':   None,
        })
    return sensors


# ── Coordinate-based matching ─────────────────────────────────────────────────

# Maximum distance to consider two sensors the same physical device.
# Sensors further apart than this are treated as unmatched.
COORD_MATCH_MAX_M = 500


def match_sensors(ref_sensors, api_sensors, max_dist=None):
    """
    Match reference sensors to API sensors by geographic proximity.

    Builds all candidate pairs within COORD_MATCH_MAX_M, sorts by distance,
    then assigns greedily shortest-first — so the globally closest pairs are
    matched before file order has any influence.

    Returns list of match dicts with type: 'match' | 'ref_only' | 'api_only'.
    """
    api_with_coords = [s for s in api_sensors if s['lat'] is not None]
    refs_with_coords = [r for r in ref_sensors if r['lat'] is not None]
    refs_no_coords   = [r for r in ref_sensors if r['lat'] is None]

    threshold = max_dist if max_dist is not None else COORD_MATCH_MAX_M

    # Build all pairs within threshold, sorted by distance. Rows are tracked by
    # position, not by name: two distinct sites can legitimately share a name
    # (e.g. two points on Georgiou Griva Digeni Avenue), and keying on the name
    # would let the first match swallow the second row's identity.
    candidates = []
    for ri, ref in enumerate(refs_with_coords):
        for api in api_with_coords:
            dist = _haversine_m(ref['lat'], ref['lon'], api['lat'], api['lon'])
            if dist <= threshold:
                candidates.append((dist, ri, api))
    candidates.sort(key=lambda x: (x[0], x[1]))

    matched_ref_idx  = set()
    matched_api_ids  = set()
    results = []

    for dist, ri, api in candidates:
        if ri in matched_ref_idx or api['id'] in matched_api_ids:
            continue
        matched_ref_idx.add(ri)
        matched_api_ids.add(api['id'])
        results.append({
            'type':       'match',
            'confidence': None,
            'ref':        refs_with_coords[ri],
            'api':        api,
            'distance_m': round(dist, 1),
        })

    for ri, ref in enumerate(refs_with_coords):
        if ri not in matched_ref_idx:
            # Find nearest API sensor for the note, even if beyond threshold
            nearest = min(api_with_coords,
                          key=lambda a: _haversine_m(ref['lat'], ref['lon'], a['lat'], a['lon']),
                          default=None)
            note = (f'Nearest API sensor is {_haversine_m(ref["lat"], ref["lon"], nearest["lat"], nearest["lon"]):.0f} m away'
                    if nearest else 'No API sensors with coordinates')
            results.append({'type': 'ref_only', 'ref': ref, 'api': None, 'note': note})

    for ref in refs_no_coords:
        results.append({'type': 'ref_only', 'ref': ref, 'api': None,
                        'note': 'No coordinates in reference file'})

    # Detect co-located pairs: unmatched API sensors within COLOCATION_M of a matched
    # one. Same radius annotate_accountability() uses to inherit a project, so the QA
    # report can never label a sensor "api_only" while it silently inherits ownership.
    matched_api_list = [r['api'] for r in results if r['type'] == 'match']
    for api in api_sensors:
        if api['id'] in matched_api_ids:
            continue  # already represented as a 'match'
        if api['lat'] is None:
            results.append({'type': 'api_only', 'ref': None, 'api': api})
            continue
        sibling = next(
            (m for m in matched_api_list
             if _haversine_m(api['lat'], api['lon'], m['lat'], m['lon']) <= COLOCATION_M),
            None
        )
        if sibling:
            results.append({'type': 'colocated', 'api': api, 'sibling': sibling})
        else:
            results.append({'type': 'api_only', 'ref': None, 'api': api})

    return results


# ── HTML generation ───────────────────────────────────────────────────────────

def _badge_health(pct):
    # Thresholds shared with the main dashboard via stability.py so the QA
    # report and the public dashboard never disagree on the same percentage.
    if pct is None:
        return '<span class="badge grey">No data</span>'
    if pct >= HEALTH_GOOD_PCT:
        return f'<span class="badge green">{pct}%</span>'
    if pct >= HEALTH_WARNING_PCT:
        return f'<span class="badge amber">{pct}%</span>'
    return f'<span class="badge red">{pct}%</span>'


def _badge_last_seen(last_seen):
    if not last_seen:
        return '<span class="badge grey">Never</span>'
    try:
        from datetime import timezone
        ts = datetime.fromisoformat(last_seen.replace('Z', '+00:00'))
        age_h = (datetime.now(timezone.utc) - ts).total_seconds() / 3600
        label = f'{int(age_h)}h ago' if age_h >= 1 else f'{int(age_h * 60)}m ago'
        if age_h < 1:
            return f'<span class="badge green">{label}</span>'
        if age_h < 24:
            return f'<span class="badge amber">{label}</span>'
        return f'<span class="badge red">{label}</span>'
    except Exception:
        return f'<span class="badge grey">{last_seen[:10]}</span>'


def _badge_dist(dist):
    if dist is None:
        return '<span class="badge grey">No ref coords</span>'
    if dist < COORD_WARN_M:
        return f'<span class="badge green">{dist:.0f} m</span>'
    if dist < COORD_ALERT_M:
        return f'<span class="badge amber">{dist:.0f} m ⚠</span>'
    return f'<span class="badge red">{dist:.0f} m ✖</span>'


def _badge_conf(score):
    if score >= 0.90:
        return f'<span class="badge green">{score:.0%}</span>'
    if score >= 0.75:
        return f'<span class="badge amber">{score:.0%} ⚠</span>'
    return f'<span class="badge red">{score:.0%} ✖</span>'


# Same wording as report.py's _COMMISSIONING_LABEL, so a row reads the same
# whether it's seen on the public dashboard or in this diagnostic tool.
_REF_STATUS_LABEL = {
    'active':          ('green', 'Active'),
    'not_electrified': ('grey',  'Awaiting power'),
    'decommissioned':  ('grey',  'Decommissioned'),
}


def _badge_ref_status(commissioning):
    """Status badge for a reference-sheet row, from its own Status column.

    This is what separates a genuine gap (active, nothing in the API) from an
    expected absence (awaiting power / decommissioned) in the "In reference"
    table — without it, every unmatched row looks equally alarming.
    """
    colour, label = _REF_STATUS_LABEL.get(commissioning, ('grey', commissioning or 'Active'))
    return f'<span class="badge {colour}">{label}</span>'


def _badge_active(active):
    """Status badge for an API sensor: still in the live inventory feed, or
    retired from it. Distinct from _badge_ref_status — this sensor has no
    reference row, so no commissioning state exists for it."""
    return '<span class="badge green">Active</span>' if active else '<span class="badge red">Retired</span>'


def generate_html(group, api_sensors, ref_sensors, matches, out_path, live=False, not_electrified=0, max_dist=None):
    try:
        from zoneinfo import ZoneInfo
        _cy = ZoneInfo("Asia/Nicosia")
    except Exception:
        _cy = timezone.utc
    now = datetime.now(_cy).strftime('%Y-%m-%d %H:%M Cyprus time')
    matched   = [m for m in matches if m['type'] == 'match']
    ref_only  = [m for m in matches if m['type'] == 'ref_only']
    api_only  = [m for m in matches if m['type'] == 'api_only']
    colocated = [m for m in matches if m['type'] == 'colocated']

    coord_issues = [m for m in matched if m['distance_m'] and m['distance_m'] >= COORD_WARN_M]

    # This report's actual matching radius — match_sensors() never pairs anything
    # further apart than this, so a legend band beyond it can never fire and
    # would only mislead the reader into thinking farther matches are tolerated.
    match_max_m = max_dist if max_dist is not None else COORD_MATCH_MAX_M
    _amber_hi = min(match_max_m, COORD_ALERT_M)
    _show_amber = match_max_m > COORD_WARN_M
    _show_red = match_max_m > COORD_ALERT_M

    # Pre-built legend fragments for the two colour-key blocks below — kept as
    # plain strings (not inline conditionals) so neither block can show a
    # distance band this report's --max-dist makes physically unreachable.
    _key_line_amber = (
        f'<span><span style="display:inline-block;width:24px;border-top:2.5px dashed #e67e22;'
        f'vertical-align:middle"></span> {COORD_WARN_M}–{_amber_hi} m</span>'
    ) if _show_amber else ''
    _key_line_red = (
        f'<span><span style="display:inline-block;width:24px;border-top:2.5px dashed #c0392b;'
        f'vertical-align:middle"></span> &gt;{COORD_ALERT_M} m</span>'
    ) if _show_red else ''
    _corner_line_amber = (
        f'<span style="display:inline-block;width:22px;border-top:2.5px dashed #e67e22;'
        f'vertical-align:middle;margin-right:6px"></span>Match {COORD_WARN_M}–{_amber_hi} m<br>'
    ) if _show_amber else ''
    _corner_line_red = (
        f'<span style="display:inline-block;width:22px;border-top:2.5px dashed #c0392b;'
        f'vertical-align:middle;margin-right:6px"></span>Match &gt;{COORD_ALERT_M} m'
    ) if _show_red else ''

    # ── Accountability breakdown ───────────────────────────────────────────────
    # Health % is computed over supported sensors only, so a dead out-of-support
    # sensor can't drag the number down. Out-of-support sensors still reporting
    # are counted separately as bonus value.
    supported      = [s for s in api_sensors if s.get('accountability', ACCT_SUPPORTED) == ACCT_SUPPORTED]
    out_of_support = [s for s in api_sensors if s.get('accountability') == ACCT_OUT_OF_SUPPORT]
    _sup_health    = [s['health_pct'] for s in supported if s['health_pct'] is not None]
    supported_uptime = round(sum(_sup_health) / len(_sup_health)) if _sup_health else None
    oos_total = len(out_of_support)
    oos_bonus = sum(1 for s in out_of_support if (s['health_pct'] or 0) >= BONUS_HEALTH_PCT)

    # ── Map data ──────────────────────────────────────────────────────────────
    # Matches within COORD_WARN_M are "the same spot" — merge ref + API into a
    # single confirmed-match marker instead of two dots joined by a line. Matches
    # at/above that threshold keep the two-dot + line view, since the gap itself
    # is the thing worth seeing.
    near_matches_by_api = {id(m['api']): m for m in matched
                            if not (m['distance_m'] and m['distance_m'] >= COORD_WARN_M)}
    near_match_ref_ids = {id(m['ref']) for m in near_matches_by_api.values()}

    def _ref_match_fields(s):
        m = near_matches_by_api.get(id(s))
        if not m:
            return {'ref_name': None, 'ref_source': None, 'ref_dist': None}
        return {'ref_name': m['ref']['name'], 'ref_source': m['ref']['source'],
                'ref_dist': round(m['distance_m'], 1) if m['distance_m'] else 0}

    api_map = [{'id': s['id'], 'name': s['name'], 'lat': s['lat'], 'lon': s['lon'],
                'active': s['active'], 'health': s['health_pct'],
                'accountability': s.get('accountability', ACCT_SUPPORTED),
                'project': s.get('project'),
                'project_source': s.get('project_source'),
                **_ref_match_fields(s)} for s in api_sensors]

    ref_map = [{'name': s['name'], 'lat': s['lat'], 'lon': s['lon'], 'source': s['source'],
                'project': s.get('project'),
                'accountability': s.get('accountability', ACCT_SUPPORTED)}
               for s in ref_sensors if s['lat'] is not None and id(s) not in near_match_ref_ids]

    link_map = [{'ref_lat': m['ref']['lat'], 'ref_lon': m['ref']['lon'],
                 'api_lat': m['api']['lat'], 'api_lon': m['api']['lon'],
                 'dist': round(m['distance_m']) if m['distance_m'] else None,
                 'ref_name': m['ref']['name'], 'api_name': m['api']['name'],
                 'ref_coords': f"{m['ref']['lat']:.5f}, {m['ref']['lon']:.5f}",
                 'api_coords': f"{m['api']['lat']:.5f}, {m['api']['lon']:.5f}"}
                for m in matched
                if m['ref']['lat'] is not None and m['api']['lat'] is not None
                and id(m['api']) not in near_matches_by_api]

    # ── Table rows ────────────────────────────────────────────────────────────
    def tr_matched():
        rows = sorted(matched, key=lambda x: x.get('distance_m') or 0, reverse=True)
        out = []
        for m in rows:
            ref, api = m['ref'], m['api']
            extra = '; '.join(f'{k}: {v}' for k, v in ref.get('extra', {}).items() if v)
            ref_coords = f'{ref["lat"]:.5f}, {ref["lon"]:.5f}' if ref['lat'] is not None else (ref.get('loc_raw') or '—')
            api_coords = f'{api["lat"]:.5f}, {api["lon"]:.5f}' if api['lat'] else '—'
            dist = m.get('distance_m') or 0
            if api['lat'] and ref['lat'] is not None and dist >= COORD_WARN_M:
                fly = (f"flyToBoth({ref['lat']},{ref['lon']},{api['lat']},{api['lon']},"
                       f"'{_esc(ref['name'])} (ref) vs {_esc(api['name'])} (API)')")
            elif api['lat']:
                fly = f"flyTo({api['lat']},{api['lon']},'{_esc(api['name'])} / {_esc(ref['name'])}')"
            else:
                fly = ''
            out.append(
                f'<tr class="clickable" onclick="{fly}" title="Click to show on map">'
                f'<td>&#128205; {_h(ref["name"])}</td>'
                f'<td>{_h(api["name"])}</td>'
                f'<td class="mono">{_h(api["id"])}</td>'
                f'<td>{_badge_dist(m["distance_m"])}</td>'
                f'<td class="mono dim">{ref_coords}</td>'
                f'<td class="mono dim">{api_coords}</td>'
                f'<td>{_badge_health(api["health_pct"])}</td>'
                f'<td>{_badge_active(api["active"])}</td>'
                f'<td class="dim">{_h(ref["source"])}</td>'
                f'<td class="dim">{_h(extra)}</td>'
                f'</tr>'
            )
        return ''.join(out)

    def tr_ref_only():
        out = []
        for m in ref_only:
            ref = m['ref']
            coords = (f'{ref["lat"]:.5f}, {ref["lon"]:.5f}'
                      if ref['lat'] is not None else
                      (ref.get('loc_raw') or '—'))
            notes = '; '.join(f'{k}: {v}' for k, v in ref.get('extra', {}).items() if v)
            note  = m.get('note', '')
            fly   = f"flyTo({ref['lat']},{ref['lon']},'{_esc(ref['name'])}')" if ref['lat'] else ''
            out.append(
                f'<tr class="clickable" onclick="{fly}" title="Click to show on map">'
                f'<td>&#128205; {_h(ref["name"])}</td>'
                f'<td class="mono">{coords}</td>'
                f'<td class="dim">{_h(ref["source"])}</td>'
                f'<td>{_badge_ref_status(ref.get("commissioning", "active"))}</td>'
                f'<td class="dim">{_h(notes)}</td>'
                f'<td class="dim">{_h(note)}</td>'
                f'</tr>'
            )
        return ''.join(out)

    def _esc(s):
        # JS string-literal escaping (for onclick handlers).
        return s.replace('\n', ' ').replace('\r', '').replace("'", "\\'") if s else ''

    def _h(s):
        # HTML text escaping for spreadsheet-/API-derived free text in table cells.
        return html.escape(str(s)) if s is not None else ''

    def _fly_link(api):
        if not api['lat']:
            return _h(api['name'])
        js = f"flyTo({api['lat']},{api['lon']},'{_esc(api['name'])}')"
        return f'<span class="fly-link" onclick="{js}" title="Show on map">&#128205; {_h(api["name"])}</span>'

    def tr_api_only():
        out = []
        for m in api_only:
            api = m['api']
            coords = f'{api["lat"]:.5f}, {api["lon"]:.5f}' if api['lat'] else '—'
            out.append(
                f'<tr>'
                f'<td>{_fly_link(api)}</td>'
                f'<td class="mono">{api["id"]}</td>'
                f'<td class="mono">{coords}</td>'
                f'<td>{_badge_health(api["health_pct"])}</td>'
                f'<td>{_badge_active(api["active"])}</td>'
                f'</tr>'
            )
        return ''.join(out)

    def tr_colocated():
        out = []
        for m in colocated:
            api, sib = m['api'], m['sibling']
            coords = f'{api["lat"]:.5f}, {api["lon"]:.5f}' if api['lat'] else '—'
            out.append(
                f'<tr>'
                f'<td>{_fly_link(api)}</td>'
                f'<td class="mono">{api["id"]}</td>'
                f'<td class="mono">{coords}</td>'
                f'<td>{_badge_health(api["health_pct"])}</td>'
                f'<td class="dim">Co-located with <b>{_h(sib["name"])}</b> (ID {_h(sib["id"])})</td>'
                f'</tr>'
            )
        return ''.join(out)

    # ── Render ────────────────────────────────────────────────────────────────
    # Not named `html`: that would shadow the module for the nested _h() escaper.
    page = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>QA Report — {group}</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:Arial,sans-serif;font-size:13px;background:#f4f6f9;color:#222}}
header{{background:#1f4e79;color:#fff;padding:16px 24px}}
header h1{{font-size:20px;font-weight:bold}}
header p{{font-size:12px;opacity:.8;margin-top:4px}}
.summary{{display:flex;gap:12px;padding:16px 24px;flex-wrap:wrap}}
.card{{background:#fff;border-radius:6px;padding:14px 20px;min-width:130px;
       box-shadow:0 1px 4px rgba(0,0,0,.1);text-align:center;border-top:3px solid #ccc}}
.card .num{{font-size:26px;font-weight:bold}}
.card .lbl{{font-size:11px;color:#666;margin-top:3px}}
.card.blue{{border-color:#1f4e79}}.card.blue .num{{color:#1f4e79}}
.card.green{{border-color:#27ae60}}.card.green .num{{color:#27ae60}}
.card.amber{{border-color:#e67e22}}.card.amber .num{{color:#e67e22}}
.card.red{{border-color:#c0392b}}.card.red .num{{color:#c0392b}}
.map-wrap{{position:sticky;top:0;z-index:900;background:#f4f6f9;padding-bottom:8px}}
#map{{height:420px;margin:0 24px 0;border-radius:6px;
      box-shadow:0 1px 6px rgba(0,0,0,.15)}}
.legend{{background:#fff;padding:10px 14px;border-radius:5px;
         line-height:2;font-size:12px;box-shadow:0 1px 4px rgba(0,0,0,.2)}}
.legend i{{display:inline-block;width:12px;height:12px;border-radius:50%;
           margin-right:6px;vertical-align:middle}}
section{{margin:0 24px 24px}}
h2{{font-size:14px;font-weight:bold;color:#1f4e79;padding-bottom:6px;
    border-bottom:2px solid #1f4e79;margin-bottom:10px}}
details{{background:#fff;border-radius:6px;box-shadow:0 1px 4px rgba(0,0,0,.08);
         overflow:hidden}}
.tbl-scroll{{max-height:320px;overflow-y:auto}}
summary{{padding:10px 14px;cursor:pointer;font-weight:bold;font-size:13px;
         list-style:none;user-select:none;background:#f8f9fa}}
summary::-webkit-details-marker{{display:none}}
summary::before{{content:'▶  '}}
details[open] summary::before{{content:'▼  '}}
table{{width:100%;border-collapse:collapse;font-size:12px}}
th{{background:#1f4e79;color:#fff;padding:7px 10px;text-align:left;white-space:nowrap;
    position:sticky;top:0;z-index:1}}
td{{padding:6px 10px;border-bottom:1px solid #eee;vertical-align:middle}}
tr:last-child td{{border-bottom:none}}
tr:hover td{{background:#f0f6ff}}
.badge{{display:inline-block;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:bold}}
.badge.green{{background:#d4edda;color:#155724}}
.badge.amber{{background:#fff3cd;color:#856404}}
.badge.red{{background:#f8d7da;color:#721c24}}
.badge.grey{{background:#e9ecef;color:#6c757d}}
.mono{{font-family:monospace;font-size:11px}}
.dim{{color:#888}}
tr.clickable{{cursor:pointer}}
tr.clickable:hover td{{background:#dceeff}}
.fly-link{{cursor:pointer;color:#1f4e79;text-decoration:underline dotted}}
.fly-link:hover{{color:#c0392b}}
.colour-key{{display:flex;gap:16px;align-items:center;padding:8px 24px 12px;flex-wrap:wrap;font-size:12px}}
.colour-key span{{display:flex;align-items:center;gap:6px}}
.colour-key i{{display:inline-block;width:14px;height:14px;border-radius:50%;border:2px solid #fff;
               box-shadow:0 0 3px rgba(0,0,0,.3);flex-shrink:0}}
.search-bar{{margin:0 24px 12px;display:flex;gap:8px;align-items:center}}
.search-bar input{{flex:1;padding:7px 12px;border:1px solid #ccc;border-radius:4px;font-size:13px}}
.search-bar label{{font-size:12px;color:#555;white-space:nowrap}}
#pin-bar{{position:absolute;bottom:12px;left:50%;transform:translateX(-50%);z-index:1000;
          background:#fff;border-radius:6px;box-shadow:0 2px 8px rgba(0,0,0,.3);
          padding:8px 14px;display:flex;align-items:center;gap:10px;font-size:13px;
          opacity:0;transition:opacity .2s;pointer-events:none}}
#pin-bar.visible{{opacity:1;pointer-events:auto}}
#pin-coords{{font-family:monospace;font-weight:bold;color:#1f4e79}}
#pin-copy{{padding:4px 10px;background:#1f4e79;color:#fff;border:none;border-radius:4px;
           cursor:pointer;font-size:12px}}
#pin-copy:hover{{background:#163a5f}}
#pin-copied{{font-size:11px;color:#27ae60;display:none}}
#ruler-bar{{position:absolute;bottom:12px;left:50%;transform:translateX(-50%);z-index:1000;
            background:#fff;border-radius:6px;box-shadow:0 2px 8px rgba(0,0,0,.3);
            padding:8px 14px;display:flex;align-items:center;gap:10px;font-size:13px;
            opacity:0;transition:opacity .2s;pointer-events:none}}
#ruler-bar.visible{{opacity:1;pointer-events:auto}}
#ruler-dist{{font-family:monospace;font-weight:bold;color:#8e44ad}}
#ruler-clear{{padding:4px 10px;background:#8e44ad;color:#fff;border:none;border-radius:4px;
              cursor:pointer;font-size:12px}}
#ruler-clear:hover{{background:#6c3483}}
</style>
</head>
<body>
<header>
  <h1>QA Report — {group}</h1>
  <p>Generated {now}&nbsp;&nbsp;|&nbsp;&nbsp;
     <b>{'&#x1F4F6; Live from API' if live else '&#x1F4BE; From local DB (cached)'}</b>&nbsp;&nbsp;|&nbsp;&nbsp;
     {len(api_sensors)} sensors in API&nbsp;&nbsp;|&nbsp;&nbsp;
     {len(ref_sensors)} sensors in reference files</p>
</header>

<div class="summary">
  <div class="card blue"><div class="num">{len(api_sensors)}</div><div class="lbl">API sensors</div></div>
  <div class="card blue"><div class="num">{len(ref_sensors)}</div><div class="lbl">Reference sensors</div></div>
  <div class="card green"><div class="num">{len(matched)}</div><div class="lbl">Matched by location</div></div>
  <div class="card amber"><div class="num">{len(coord_issues)}</div><div class="lbl">Coord mismatches (&gt;{COORD_WARN_M}m)</div></div>
  <div class="card red"><div class="num">{len(ref_only)}</div><div class="lbl">In ref, not in API</div></div>
  <div class="card red"><div class="num">{len(api_only)}</div><div class="lbl">In API, not in ref</div></div>
  {'<div class="card amber"><div class="num">'+str(len(colocated))+'</div><div class="lbl">Co-located pairs</div></div>' if colocated else ''}
  {'<div class="card green"><div class="num">'+str(supported_uptime)+'%</div><div class="lbl">Supported uptime</div></div>' if supported_uptime is not None else ''}
  {'<div class="card grey"><div class="num">'+str(oos_total)+'</div><div class="lbl">Out of support (★ '+str(oos_bonus)+' still reporting)</div></div>' if oos_total else ''}
  {'<div class="card grey"><div class="num">'+str(not_electrified)+'</div><div class="lbl">Awaiting power (excluded from stats)</div></div>' if not_electrified else ''}
</div>

<div class="map-wrap">
<div class="colour-key">
  <strong>Map dots:</strong>
  <span><i style="background:#27ae60"></i> API — healthy (&ge;90% uptime)</span>
  <span><i style="background:#e67e22"></i> API — degraded (70–89%)</span>
  <span><i style="background:#c0392b"></i> API — critical (&lt;70%)</span>
  <span><i style="background:#888"></i> API — retired or no history</span>
  <span><i style="background:#fff;border:3px solid #27ae60;box-shadow:none"></i> Out of support — still reporting (bonus, &ge;{BONUS_HEALTH_PCT}% uptime)</span>
  <span><i style="background:#fff;border:3px solid #95a5a6;box-shadow:none"></i> Out of support — offline / occasional (&lt;{BONUS_HEALTH_PCT}%)</span>
  <span><i style="background:#1f4e79"></i> Your spreadsheet location</span>
  <span><i style="background:#fff;border:3px solid #1f4e79;box-shadow:none"></i> Spreadsheet location — out of support</span>
  <span><i style="background:#27ae60;outline:3px solid #f1c40f;outline-offset:2px"></i> Confirmed match — ref + API agree (&lt;{COORD_WARN_M} m, shown as one marker)</span>
  <strong style="margin-left:8px">Lines:</strong>
  <span><span style="display:inline-block;width:24px;border-top:2.5px dashed #27ae60;vertical-align:middle"></span> &lt;{COORD_WARN_M} m apart</span>
  {_key_line_amber}
  {_key_line_red}
</div>
<div style="padding:0 24px 10px;font-size:11px;color:#888">
  <b>Note:</b> Out-of-support sensors belong to a project whose maintenance contract has ended — a failure is
  expected, not actionable. They are excluded from the <b>Supported uptime</b> figure. One counts as
  <b>bonus</b> (★) only if it still delivers <b>&ge;{BONUS_HEALTH_PCT}% uptime</b>; between 0 and {BONUS_HEALTH_PCT}% it is
  merely <i>occasionally reporting</i>, and at 0% it is offline (all expected).
</div>

<div style="position:relative">
  <div id="map"></div>
  <div id="pin-toggle" onclick="togglePinMode()"
       style="position:absolute;top:10px;right:10px;z-index:1000;
              background:#fff;border-radius:6px;box-shadow:0 2px 6px rgba(0,0,0,.3);
              padding:6px 12px;cursor:pointer;font-size:12px;user-select:none;
              border:2px solid transparent">
    &#128205; Get coordinates
  </div>
  <div id="ruler-toggle" onclick="toggleRulerMode()"
       style="position:absolute;top:48px;right:10px;z-index:1000;
              background:#fff;border-radius:6px;box-shadow:0 2px 6px rgba(0,0,0,.3);
              padding:6px 12px;cursor:pointer;font-size:12px;user-select:none;
              border:2px solid transparent">
    &#128207; Ruler
  </div>
  <div id="pin-bar">
    <span>&#128205;</span>
    <span id="pin-coords"></span>
    <button id="pin-copy" onclick="pinCopy()">Copy</button>
    <span id="pin-copied">Copied!</span>
  </div>
  <div id="ruler-bar">
    <span>&#128207;</span>
    <span id="ruler-dist">Click two points on the map</span>
    <button id="ruler-clear" onclick="rulerClear()">Clear</button>
  </div>
</div>
</div><!-- /map-wrap -->

<div class="search-bar">
  <label>Search all tables:</label>
  <input type="text" id="search" placeholder="Type a name, ID, or description..." oninput="filterTables(this.value)">
</div>

<section>
  <h2>Matched sensors ({len(matched)})</h2>
  <details open>
    <summary>Sorted by coordinate distance — largest first</summary>
    <div class="tbl-scroll">
    <table>
      <thead><tr>
        <th>Reference name</th><th>API description</th><th>API ID</th>
        <th>Coord distance</th><th>Ref GPS (spreadsheet)</th><th>API GPS</th>
        <th>Health</th><th>Status</th><th>Source file</th><th>Notes</th>
      </tr></thead>
      <tbody>{tr_matched()}</tbody>
    </table>
    </div>
  </details>
</section>

<section>
  <h2>In reference — not found in API ({len(ref_only)})</h2>
  <details {'open' if ref_only else ''}>
    <summary>Sensors in your spreadsheets with no API match — not installed, wrong ID, or data entry error</summary>
    <div class="tbl-scroll">
    <table>
      <thead><tr><th>Name</th><th>Coordinates</th><th>Source file</th><th>Status</th><th>Notes</th><th>Reason</th></tr></thead>
      <tbody>{tr_ref_only()}</tbody>
    </table>
    </div>
  </details>
</section>

<section>
  <h2>In API — not in reference ({len(api_only)})</h2>
  <details {'open' if api_only else ''}>
    <summary>Sensors the API reports that are not in any spreadsheet — undocumented or recently added</summary>
    <div class="tbl-scroll">
    <table>
      <thead><tr><th>API name</th><th>API ID</th><th>Coordinates</th><th>Health</th><th>Status</th></tr></thead>
      <tbody>{tr_api_only()}</tbody>
    </table>
    </div>
  </details>
</section>

{'<section>' if colocated else ''}
{f'<h2>&#128101; Co-located Pairs ({len(colocated)})</h2>' if colocated else ''}
{'<p class="dim">These API sensors share the same physical location as a matched sensor. They likely represent the opposite direction of travel. Add a second row in your spreadsheet for each one.</p>' if colocated else ''}
{'''<details open>
  <summary>API sensors co-located with a matched sensor (&le; 10 m apart) — not yet in any spreadsheet</summary>
  <div class="tbl-scroll">
  <table>
    <thead><tr><th>API name</th><th>API ID</th><th>Coordinates</th><th>Health</th><th>Note</th></tr></thead>
    <tbody>''' + tr_colocated() + '''</tbody>
  </table>
  </div>
</details>
</section>''' if colocated else ''}

<script>
const apiMarkers = {json.dumps(api_map)};
const refMarkers = {json.dumps(ref_map)};
const linkLines  = {json.dumps(link_map)};
const WARN_M = {COORD_WARN_M}, ALERT_M = {COORD_ALERT_M};

const map = L.map('map');
L.tileLayer('https://{{s}}.tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png',
  {{attribution:'© OpenStreetMap contributors', maxZoom:19}}).addTo(map);

function dot(color, size=14, hollow=false, confirmed=false) {{
  // hollow = out-of-support: transparent fill + thick coloured ring, so these
  // read as a different *class* of marker, not just a different health colour.
  // confirmed = reference and API agree on this spot — draw a gold halo so a
  // matched pair reads as one unit at a glance, without needing two dots + a line.
  const inner = hollow
    ? `background:#fff;border:3px solid ${{color}};box-shadow:0 0 4px rgba(0,0,0,.35)`
    : `background:${{color}};border:2px solid #fff;box-shadow:0 0 4px rgba(0,0,0,.4)`;
  const halo = confirmed ? `outline:3px solid #f1c40f;outline-offset:2px;` : '';
  return L.divIcon({{
    className:'',
    html:`<div style="width:${{size}}px;height:${{size}}px;border-radius:50%;${{inner}};${{halo}}"></div>`,
    iconSize:[size,size], iconAnchor:[size/2,size/2]
  }});
}}

const bounds = [];

// Group API markers by coordinate — sensors sharing a location form a cluster
const apiGroups = {{}};
apiMarkers.forEach(s => {{
  if (s.lat == null) return;
  const key = s.lat.toFixed(5) + ',' + s.lon.toFixed(5);
  if (!apiGroups[key]) apiGroups[key] = [];
  apiGroups[key].push(s);
}});

Object.values(apiGroups).forEach(group => {{
  const lat = group[0].lat, lon = group[0].lon;
  const n = group.length;

  // Colour driven by worst health in the group
  function sensorColor(s) {{
    return !s.active ? '#888' : s.health == null ? '#aaa'
         : s.health >= 90 ? '#27ae60' : s.health >= 70 ? '#e67e22' : '#c0392b';
  }}
  // A group is out-of-support only if EVERY sensor in it is; a mixed cluster
  // keeps normal colouring so any actionable (supported) fault stays visible.
  const allOOS = group.every(s => s.accountability === 'out_of_support');
  let color, hollow = false;
  if (allOOS) {{
    hollow = true;
    // green ring if any is still reporting (bonus value), grey if all dark
    color = group.some(s => (s.health || 0) >= {BONUS_HEALTH_PCT}) ? '#27ae60' : '#95a5a6';
  }} else {{
    const colors = group.map(sensorColor);
    const priority = ['#c0392b','#e67e22','#27ae60','#aaa','#888'];
    color = priority.find(c => colors.includes(c)) || '#aaa';
  }}

  // Confirmed if any sensor in the group has a matched reference — the whole
  // cluster gets the gold halo since they share one map position.
  const confirmed = group.some(s => s.ref_name);

  const ringInner = hollow
    ? `background:#fff;border:3px solid ${{color}};color:${{color}}`
    : `background:${{color}};border:2px solid #fff;color:#fff`;
  const halo = confirmed ? `outline:3px solid #f1c40f;outline-offset:2px;` : '';
  const icon = n === 1
    ? dot(color, 16, hollow, confirmed)
    : L.divIcon({{
        className: '',
        html: `<div style="width:24px;height:24px;border-radius:50%;${{ringInner}};${{halo}}
                    box-shadow:0 0 4px rgba(0,0,0,.4);
                    display:flex;align-items:center;justify-content:center;
                    font-size:11px;font-weight:bold">${{n}}</div>`,
        iconSize: [24,24], iconAnchor: [12,12]
      }});

  const popup = group.map(s => {{
    const projectLine = s.project
      ? `<br>Project: <b>${{s.project}}</b>`
        + (s.project_source === 'colocated'
            ? ` <span style="color:#aaa">(inherited from co-located sensor)</span>` : '')
      : `<br><span style="color:#aaa">Project: not in reference</span>`;
    const oosNote = (s.health || 0) >= {BONUS_HEALTH_PCT}
      ? ' — ★ still reporting (bonus)'
      : ((s.health || 0) > 0 ? ' — occasionally reporting' : ' — offline (expected)');
    const acct = s.accountability === 'out_of_support'
      ? `<br><span style="color:#7f8c8d">Out of support` + `${{oosNote}}</span>`
      : '';
    const refLine = s.ref_name
      ? `<br><span style="color:#b8860b">&#10003; Confirmed match: <b>${{s.ref_name}}</b></span>`
        + (s.ref_source ? ` <span style="color:#aaa">(${{s.ref_source}})</span>` : '')
        + (s.ref_dist ? ` <span style="color:#aaa">${{s.ref_dist}} m</span>` : '')
      : '';
    return `<b>API: ${{s.name}}</b><br>ID: ${{s.id}}<br>`+
      `Health: ${{s.health != null ? s.health+'%' : 'No data'}}<br>`+
      `Active: ${{s.active ? 'Yes' : 'No'}}` + projectLine + acct + refLine;
  }}).join('<hr style="margin:6px 0">');

  L.marker([lat, lon], {{icon}}).bindPopup(popup).addTo(map);
  bounds.push([lat, lon]);
}});

refMarkers.forEach(s => {{
  const oos = s.accountability === 'out_of_support';
  const projectLine = s.project
    ? `<br>Project: <b>${{s.project}}</b>`
    : `<br><span style="color:#aaa">Project: none listed</span>`;
  const acctLine = oos
    ? `<br><span style="color:#7f8c8d">Out of support — gap here is expected</span>` : '';
  L.marker([s.lat, s.lon], {{icon: dot('#1f4e79', 14, oos)}})
   .bindPopup(`<b>REF: ${{s.name}}</b><br>Source: ${{s.source}}` + projectLine + acctLine)
   .addTo(map);
  bounds.push([s.lat, s.lon]);
}});

linkLines.forEach(l => {{
  const color = l.dist == null ? '#aaa'
              : l.dist < WARN_M ? '#27ae60'
              : l.dist < ALERT_M ? '#e67e22' : '#c0392b';
  L.polyline([[l.ref_lat,l.ref_lon],[l.api_lat,l.api_lon]],
    {{color, weight:2.5, opacity:.85, dashArray:'6,4'}})
   .bindPopup(
     `<b>${{l.ref_name}}</b> ↔ <b>${{l.api_name}}</b><br>`+
     `<span style="color:#555;font-size:11px">`+
     `&#128205; Spreadsheet: ${{l.ref_coords}}<br>`+
     `&#x1F4F6; API: ${{l.api_coords}}<br>`+
     `Distance: ${{l.dist != null ? '<b>'+l.dist+' m</b>' : 'unknown'}}</span>`)
   .addTo(map);
}});

const legend = L.control({{position:'bottomright'}});
legend.onAdd = () => {{
  const d = L.DomUtil.create('div','legend');
  d.innerHTML =
    '<b>Legend</b><br>'+
    '<i style="background:#27ae60"></i>API healthy (&ge;90%)<br>'+
    '<i style="background:#e67e22"></i>API degraded (70–90%)<br>'+
    '<i style="background:#c0392b"></i>API critical (&lt;70%)<br>'+
    '<i style="background:#888"></i>API retired / no data<br>'+
    '<i style="background:#1f4e79"></i>Reference location<br>'+
    '<i style="background:#27ae60;outline:3px solid #f1c40f;outline-offset:2px"></i>Confirmed match (ref + API agree, &lt;{COORD_WARN_M} m)<br>'+
    '<span style="display:inline-block;width:22px;border-top:2.5px dashed #27ae60;vertical-align:middle;margin-right:6px"></span>Match &lt;{COORD_WARN_M} m<br>'+
    '{_corner_line_amber}'+
    '{_corner_line_red}';
  return d;
}};
legend.addTo(map);

if (bounds.length) map.fitBounds(bounds, {{padding:[40,40]}});
else map.setView([34.9,33.0],9);

// ── Fly to sensor when clicking a table row ───────────────────────────────
let activePopup = null;
function flyTo(lat, lon, label) {{
  // #map is position:sticky, so it's already pinned in view while scrolling
  // through the tables below it — no page scroll needed, just move the camera.
  map.flyTo([lat, lon], 16, {{duration: 1.2}});
  if (activePopup) activePopup.remove();
  activePopup = L.popup()
    .setLatLng([lat, lon])
    .setContent(`<b>${{label}}</b>`)
    .openOn(map);
}}

function flyToBoth(refLat, refLon, apiLat, apiLon, label) {{
  // #map is position:sticky, so it's already pinned in view while scrolling
  // through the tables below it — no page scroll needed, just move the camera.
  const b = L.latLngBounds([[refLat, refLon], [apiLat, apiLon]]);
  map.flyToBounds(b, {{padding:[80,80], maxZoom:17, duration:1.2}});
  if (activePopup) activePopup.remove();
  const midLat = (refLat + apiLat) / 2, midLon = (refLon + apiLon) / 2;
  activePopup = L.popup()
    .setLatLng([midLat, midLon])
    .setContent(`<b>${{label}}</b><br><span style="font-size:11px;color:#555">Blue dot = spreadsheet &nbsp;|&nbsp; Coloured dot = API</span>`)
    .openOn(map);
}}

// ── Click-to-pin coordinates ─────────────────────────────────────────────
let pinMarker = null;
let pinCoords = '';
let pinMode = false;

function togglePinMode() {{
  pinMode = !pinMode;
  const btn = document.getElementById('pin-toggle');
  if (pinMode) {{
    if (rulerMode) toggleRulerMode();
    btn.style.borderColor = '#1f4e79';
    btn.style.background = '#e8f0fa';
    map.getContainer().style.cursor = 'crosshair';
  }} else {{
    btn.style.borderColor = 'transparent';
    btn.style.background = '#fff';
    map.getContainer().style.cursor = '';
    document.getElementById('pin-bar').classList.remove('visible');
    if (pinMarker) {{ map.removeLayer(pinMarker); pinMarker = null; }}
  }}
}}

function pinCopy() {{
  navigator.clipboard.writeText(pinCoords).then(() => {{
    const el = document.getElementById('pin-copied');
    el.style.display = '';
    setTimeout(() => el.style.display = 'none', 2000);
  }});
}}

// ── Click-to-measure ruler ────────────────────────────────────────────────
// Same haversine formula as _haversine_m() in qa.py, so the on-map distance
// always matches what the matcher itself would compute for those two points.
let rulerMode = false;
let rulerPoints = [];
let rulerMarkers = [];
let rulerLine = null;

function haversineM(lat1, lon1, lat2, lon2) {{
  const R = 6371000, rad = Math.PI / 180;
  const phi1 = lat1 * rad, phi2 = lat2 * rad;
  const dphi = (lat2 - lat1) * rad, dlam = (lon2 - lon1) * rad;
  const a = Math.sin(dphi/2)**2 + Math.cos(phi1) * Math.cos(phi2) * Math.sin(dlam/2)**2;
  return R * 2 * Math.atan2(Math.sqrt(a), Math.sqrt(1 - a));
}}

function fmtDist(m) {{
  return m >= 1000 ? `${{(m/1000).toFixed(2)}} km` : `${{Math.round(m)}} m`;
}}

function toggleRulerMode() {{
  rulerMode = !rulerMode;
  const btn = document.getElementById('ruler-toggle');
  if (rulerMode) {{
    if (pinMode) togglePinMode();
    btn.style.borderColor = '#8e44ad';
    btn.style.background = '#f3e8fa';
    map.getContainer().style.cursor = 'crosshair';
    document.getElementById('ruler-bar').classList.add('visible');
  }} else {{
    btn.style.borderColor = 'transparent';
    btn.style.background = '#fff';
    map.getContainer().style.cursor = '';
    document.getElementById('ruler-bar').classList.remove('visible');
    rulerClear();
  }}
}}

function rulerClear() {{
  rulerPoints = [];
  rulerMarkers.forEach(m => map.removeLayer(m));
  rulerMarkers = [];
  if (rulerLine) {{ map.removeLayer(rulerLine); rulerLine = null; }}
  document.getElementById('ruler-dist').textContent = 'Click two points on the map';
}}

map.on('click', function(e) {{
  if (pinMode) {{
    const lat = e.latlng.lat.toFixed(6);
    const lon = e.latlng.lng.toFixed(6);
    pinCoords = `${{lat}}, ${{lon}}`;
    if (pinMarker) pinMarker.setLatLng(e.latlng);
    else pinMarker = L.marker(e.latlng).addTo(map);
    document.getElementById('pin-coords').textContent = pinCoords;
    document.getElementById('pin-bar').classList.add('visible');
    document.getElementById('pin-copied').style.display = 'none';
    return;
  }}
  if (rulerMode) {{
    if (rulerPoints.length >= 2) rulerClear();  // 3rd click starts a fresh measurement
    rulerPoints.push(e.latlng);
    rulerMarkers.push(L.circleMarker(e.latlng, {{radius:5, color:'#8e44ad', fillColor:'#8e44ad', fillOpacity:1}}).addTo(map));
    if (rulerPoints.length === 1) {{
      document.getElementById('ruler-dist').textContent = 'Click a second point…';
    }} else {{
      rulerLine = L.polyline(rulerPoints, {{color:'#8e44ad', weight:3, dashArray:'6,6'}}).addTo(map);
      const d = haversineM(rulerPoints[0].lat, rulerPoints[0].lng, rulerPoints[1].lat, rulerPoints[1].lng);
      document.getElementById('ruler-dist').textContent = fmtDist(d);
    }}
  }}
}});

// ── Search / filter all table rows ───────────────────────────────────────
function filterTables(query) {{
  const q = query.toLowerCase().trim();
  document.querySelectorAll('tbody tr').forEach(row => {{
    const text = row.textContent.toLowerCase();
    row.style.display = (!q || text.includes(q)) ? '' : 'none';
  }});
}}
</script>
</body>
</html>"""

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(page, encoding='utf-8')
    return out_path


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="ITS QA — sensor data quality report",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument('--group', required=True,
                        help='Sensor group name as stored in DB (e.g. VMS, "Traffic Detection")')
    parser.add_argument('--ref', nargs='*', default=[], metavar='FILE',
                        help='Reference CSV or Excel files (optional)')
    parser.add_argument('--out', default=None,
                        help='Output HTML path (default: reports/qa_<group>.html)')
    parser.add_argument('--max-dist', type=int, default=None,
                        help=f'Max match distance in metres (default: {COORD_MATCH_MAX_M})')
    args = parser.parse_args()

    safe_name = re.sub(r'[^a-z0-9]+', '_', args.group.lower()).strip('_')
    out_path  = Path(args.out) if args.out else REPORT_DIR / f'qa_{safe_name}.html'

    live_mode = os.environ.get('LIVE_MODE', '').lower() in ('1', 'true', 'yes')

    print(f"\nGroup: {args.group}")
    if live_mode:
        print("Loading API sensors live (LIVE_MODE=true) …")
        api_sensors = load_api_sensors_live(args.group)
    else:
        print("Loading API sensors from DB …")
        api_sensors = load_api_sensors(args.group)
    print(f"  ->{len(api_sensors)} sensors found")

    ref_sensors = []
    not_electrified = 0
    if args.ref:
        print(f"Loading reference files …")
        ref_sensors, not_electrified = load_reference(args.ref)
        print(f"  ->{len(ref_sensors)} sensors total across {len(args.ref)} file(s)")
        if not_electrified:
            print(f"  {not_electrified} sensor(s) awaiting power — flagged, excluded from health stats")

    matches = match_sensors(ref_sensors, api_sensors, max_dist=args.max_dist) if ref_sensors else []

    # Accountability: API sensors inherit it via the matched reference row;
    # reference sensors get it directly from their own Project column.
    project_acct = load_project_accountability()
    annotate_accountability(api_sensors, matches, project_acct, max_dist=args.max_dist)
    annotate_ref_accountability(ref_sensors, project_acct)
    _oos = sum(1 for s in api_sensors if s['accountability'] == ACCT_OUT_OF_SUPPORT)
    if _oos:
        print(f"  {_oos} matched sensor(s) belong to an out-of-support project")

    # Persist sensor_id -> project to the DB so the regular dashboard
    # (report.py, which doesn't have the reference spreadsheet) can show
    # who owns a failing sensor without re-running this matching step.
    if ref_sensors:
        to_persist = {s['id']: {'project': s.get('project'), 'source': s.get('project_source'),
                                'commissioning': s.get('commissioning', 'active')}
                      for s in api_sensors}
        upsert_sensor_projects(args.group, to_persist)
        _known = sum(1 for v in to_persist.values() if v['project'])
        print(f"  Persisted project assignments to DB: {_known}/{len(to_persist)} sensors")

    if matches:
        matched  = sum(1 for m in matches if m['type'] == 'match')
        ref_only = sum(1 for m in matches if m['type'] == 'ref_only')
        api_only = sum(1 for m in matches if m['type'] == 'api_only')
        print(f"\nMatching results:")
        print(f"  Matched:          {matched}")
        print(f"  Ref only (gaps):  {ref_only}")
        print(f"  API only (extra): {api_only}")

    path = generate_html(args.group, api_sensors, ref_sensors, matches, out_path, live=live_mode,
                          not_electrified=not_electrified, max_dist=args.max_dist)
    print(f"\nReport written -> {path}")
    print(f"Open in browser: file:///{path.as_posix()}\n")


if __name__ == '__main__':
    main()
